#! python3

# textFileToSpreadsheet.py - Reads in the contents of several text files from desktop, and insert the contents
# into a spreadsheet; one line of text per row, one text file per column

import os
import openpyxl
from openpyxl.utils import get_column_letter
import sys

desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop\\')

# Converts the text file to spreadsheet
def textFileToSpreadsheet(listOfTextfiles):
    os.chdir(desktop)
    
    # Validates if files exist
    for textFile in listOfTextfiles:
        if not os.path.isfile(textFile) or not textFile.endswith('.txt'):
            print('One or more file list does not exist or is invalid')
            print('Please try again')
            exit()

    wb = openpyxl.Workbook()
    sheet = wb.active

    for i in range(len(listOfTextfiles)):
        # Open each .txt file
        txtFile = open(listOfTextfiles[i])

        # Read the first line of the .txt file
        line = txtFile.readline()
        
        # Row counter. keeps track of each line in the .txt file
        cnt = 1

        # While there's more lines to read
        while line:
            # Assign the line from .txt file to the row
            sheet[get_column_letter(i + 1) + str(cnt)] = line
            # Read in next line
            line = txtFile.readline()
            # Increment for next row and line
            cnt += 1

        # Close the .txt file
        txtFile.close()

    # Saves the resulting spreadsheet on desktop
    wb.save('textFileToSpreadsheet.xlsx')

# Get text files via command line
if len(sys.argv) > 1:
    textFileToSpreadsheet(sys.argv[1:])
else:
    print('Please enter the name of the text file')